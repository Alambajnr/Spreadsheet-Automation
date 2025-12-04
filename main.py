import openpyxl

# Load workbook and sheet
inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Loop through rows
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation of numbers for  products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Products with inventory < 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # write inventory value back into column 5
    product_list.cell(product_row, 5).value = inventory * price

# --- Create a summary sheet ---
if "Summary" in inv_file.sheetnames:
    summary_sheet = inv_file["Summary"]
else:
    summary_sheet = inv_file.create_sheet("Summary")

# Add headers
summary_sheet["A1"] = "Supplier Name"
summary_sheet["B1"] = "Number of Products"
summary_sheet["C1"] = "Total Inventory Value"

# Write supplier data
row_num = 2
for supplier, num_products in products_per_supplier.items():
    summary_sheet.cell(row_num, 1).value = supplier
    summary_sheet.cell(row_num, 2).value = num_products
    summary_sheet.cell(row_num, 3).value = total_value_per_supplier[supplier]
    row_num += 1

# Save workbook
inv_file.save("inventory_with_summary.xlsx")

print("Products per supplier:", products_per_supplier)
print("Total value per supplier:", total_value_per_supplier)
print("Products under 10 inventory:", products_under_10_inv)
print("✅ Inventory values saved in column 5 and summary sheet created!")